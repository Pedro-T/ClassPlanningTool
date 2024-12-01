from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, Alignment

from collections import OrderedDict

_TITLE_FONT: Font = Font(name="Helvetica", size=24)
_VALUE_FONT: Font = Font(name="Helvetica", size=11)
_VALUE_FONT_BOLD: Font = Font(name="Helvetica", bold=True, size=11)
_CENTER_ALIGNMENT: Alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
_SEMESTER_CAPACITY: int = 4

_COLUMN_WIDTHS: dict[str, int] = {
    "A": 10,
    "B": 15,
    "C": 40,
    "D": 15,
    "E": 40,
    "F": 15,
    "G": 40,
}

def write_header(ws: Worksheet) -> None:
    """
    Writes the centered title for the table

    Args:
        sheet (Worksheet): the sheet to modify
    """
    ws.merge_cells("B2:G2")
    title_cell: Cell = ws["B2"]
    title_cell.value = "Study Plan"
    title_cell.font = _TITLE_FONT
    title_cell.alignment = _CENTER_ALIGNMENT

def format_sheet(ws: Worksheet) -> None:
    """
    Sets the column widths for the sheet based on _COLUMN_WIDTHS

    Args:
        sheet (Worksheet)   : the sheet to modify
    """
    for column_letter, width in _COLUMN_WIDTHS.items():
        ws.column_dimensions[column_letter].width = width
    ws.row_dimensions[2].height = 30

def write_semester_block(ws: Worksheet, row_start: int, column_start: int, semester: str, courses: list[dict[str, str]]) -> None:
    """
    Write the block for one semester, which is a simple two column area of code | title
    Header is semester name
    Last row is a class count for the semester
    Maximum length in courses defined in _SEMESTER_CAPACITY

    Args:
        sheet (Worksheet)               : the sheet to modify
        row_start (int)                 : row offset from the top where this block starts
        column_start (int)              : column offset from the left where this block starts
        courses (list[dict[str, str]])   : list of courses to include
    """

    # semester name
    ws.merge_cells(start_row=row_start, start_column=column_start, end_row=row_start, end_column=column_start+1)
    header: Cell = ws.cell(row_start, column_start)
    header.value = semester
    header.font = _VALUE_FONT_BOLD
    header.alignment = _CENTER_ALIGNMENT

    # course list
    detail_order: tuple[str] = ("code", "title")
    fonts_order: tuple[str] = (_VALUE_FONT_BOLD, _VALUE_FONT)
    for course_idx, course in enumerate(courses):
        for col_idx_offset in range(len(detail_order)):
            cell: Cell = ws.cell(row_start+course_idx+1, column_start+col_idx_offset)
            cell.value = course[detail_order[col_idx_offset]] if course else ""
            cell.font = fonts_order[col_idx_offset]
            cell.alignment = _CENTER_ALIGNMENT
    
    # course count

    ws.merge_cells(start_row=row_start+_SEMESTER_CAPACITY+1, start_column=column_start, end_row=row_start+_SEMESTER_CAPACITY+1, end_column=column_start+1)
    footer: Cell = ws.cell(row_start+_SEMESTER_CAPACITY+1, column_start)
    footer.value = f"Courses: {len(courses)}"
    footer.font = _VALUE_FONT_BOLD
    footer.alignment = _CENTER_ALIGNMENT


def write_plan_workbook(course_plan: OrderedDict[str, list[dict[str, str]]], book_path: str) -> None:
    """
    Write the supplied study plan to an Excel sheet.

    Args:
        course_plan (OrderedDict[str, list[dict[str, str]]]): course plan to export
        book_path (str): File path where the workbook will be saved.
    
    Raises:
        Exception: If the file cannot be written.
    """
    # Debug print to verify the file path
    print(f"Attempting to write Excel file to: {book_path}")

    # Validate course plan length
    if len(course_plan.keys()) % 3 != 0:
        raise ValueError(
            f"Course plans must be in academic year sequences of three semesters. "
            f"The expected length is a multiple of three. Provided length: {len(course_plan.keys())}\n{course_plan}"
        )
    
    wb: Workbook = Workbook()
    ws: Worksheet = wb.active
    ws.title = "Course Plan"
    format_sheet(ws)
    write_header(ws)

    col_counter: int = 0
    row_index: int = ws.max_row + 1
    for semester_name, courses in course_plan.items():
        write_semester_block(ws, row_index, 2+col_counter*2, semester_name, courses)

        col_counter += 1

        if col_counter == 3:
            row_index = ws.max_row + 1
            col_counter = 0

    # Save workbook to the specified path
    try:
        wb.save(book_path)
        print(f"Excel file successfully saved at: {book_path}")
    except Exception as e:
        print(f"Failed to save Excel file: {e}")
        raise
