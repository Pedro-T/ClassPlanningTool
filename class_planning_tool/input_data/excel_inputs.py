
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from re import compile

# intends to capture any combo of F/O/D/N with optional commas or spaces, potentially followed by (May) as some of the summer columns have
COURSE_AVAILABLE_PATTERN = compile(r"^[FODN\s,]+(?:\(May\))?$") 


def get_cutoff_format(semester: str) -> int:
    """
    Convert semester string into a three digit integer to facilitate comparison. Primarily used for the cutoff
    comparison to eliminate semester data that is out of scope (too old) for planning.

    Args:
        semester (str): semester identifier in AA00 format, with one of SP SU FA seasons
    
    Returns:
        int: value to compare
    
    Raises:
        ValueError if the semester string format is not as expected

    """
    season_values: dict[str, int] = {
        "SP": 1,
        "SU": 2,
        "FA": 3
    }

    if len(semester) != 4 or semester[:2] not in season_values or not semester[2:].isdigit():
        raise ValueError("Unexpected semester format - should be SP or SU or FA followed by two digit year.")
    
    return int(semester[2:]) * 10 + season_values[semester[:2]]


def populate_column_semester_map(row_values: list[str], cutoff_input: str="") -> dict[int, str]:
    """
    Constructs a column map to more easily tag classes to the correct semester

    Args:
        row_values (str): list of strings of row values from the table header
        cutoff (str): optional cutoff value in SP24 format
    
    Returns:
        dict[int, str] of row indexes to semester string identifiers
    """
    column_map: dict[int, str] = {}
    cutoff_value: int = get_cutoff_format(cutoff_input) if cutoff_input else 0
    for idx, value in enumerate(row_values):
        if not value:
            continue
        if value[:2] not in ("SP", "SU", "FA"):
            continue
        if cutoff_value and get_cutoff_format(value) < cutoff_value:
            continue
        column_map[value] = idx
    return column_map



def extract_sheet_data(sheet: Worksheet, cutoff: str="") -> dict[str, list[str]]:
    """
    Extracts a map of course codes to semester identifiers from the course schedule sheet

    Args:
        sheet (Worksheet): openpyxl worksheet object to extract data from
        cutoff (str): optional cutoff semester value if desired

    """
    col_semester_map: dict[int, str] = {}
    results: dict[str, list[str]] = {}

    for row in sheet.iter_rows(min_row=3):
        course: str = row[0].value
        if not course:
            continue
        if course == "Course":
            col_semester_map = populate_column_semester_map([str(cell.value) for cell in row], cutoff_input=cutoff)
            continue
        row_vals: list[str] = [str(cell.value).replace("?", "") if cell else "" for cell in row]  # assumption is made that semesters marked ?? turn out to be offered
        results[course] = [
            semester_code for semester_code, index in col_semester_map.items()
              if row_vals[index] 
              and len(row_vals[index]) < 8 
              and COURSE_AVAILABLE_PATTERN.findall(row_vals[index])]
    return results


def get_class_schedule_data(file_path: str, start_semester: str="") -> dict[str, list[str]]:
    """

    Open an Excel workbook at the target path, parse the information, and return the schedule data by semester
    
    Args:
        file_path (str): path of the Excel workbook to use
        start_semester (str): optional cutoff starting semester in 'SP24' format to ignore semesters before this one

    Returns:
        dict[str, list[str]]: Dictionary representing the course listings by semester.
    
    Raises:
        FileNotFoundError if the file does not exist
        IsADirectoryError if the target path is a directory instead of a file
        InvalidFileException if the file is not an Excel workbook or is not a file
        PermissionError if the file cannot be opened due to a permissions issue
    
    """
    path: Path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"The path {file_path} does not exist.")
    if path.is_dir():
        raise IsADirectoryError(f"The path {file_path} is a directory.")

    wb: Workbook = load_workbook(Path(file_path), data_only=True)
    return extract_sheet_data(wb.active, cutoff=start_semester)  # it is assumed that the wb only has one sheet
