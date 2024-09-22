import unittest
from pathlib import Path
from class_planning_tool.input_data.excel_inputs import get_cutoff_format, populate_column_semester_map, extract_sheet_data, get_class_schedule_data
from openpyxl import load_workbook


class TestClassScheduleParsing(unittest.TestCase):

    def setUp(self):
        self.resource_path: Path = Path("./tests/resources")

    def test_semester_cutoff_formatter_valid(self):
        self.assertEqual(241, get_cutoff_format("SP24"))
        self.assertEqual(313, get_cutoff_format("FA31"))
        self.assertEqual(252, get_cutoff_format("SU25"))
    
    def test_semester_cutoff_formatter_invalid(self):
        with self.assertRaises(ValueError):
            get_cutoff_format("AA25")
        with self.assertRaises(ValueError):
            get_cutoff_format("FAA6")
        with self.assertRaises(ValueError):
            get_cutoff_format("SU5")
        
    def test_semester_index_map_valid_no_cutoff(self):
        row: list[str] = ["Course", "Course Title", "SP20", "SU20", "FA20"]
        expectation: dict[int, str] = {
            "SP20": 2,
            "SU20": 3,
            "FA20": 4
        }
        self.assertDictEqual(expectation, populate_column_semester_map(row))

    def test_semester_index_map_invalid(self):
        row: list[str] = ["Course", "Course Title", "ABCD", "NONSENSE"]
        self.assertFalse(populate_column_semester_map(row))
    

    def test_semester_index_map_valid_with_cutoff(self):
        row: list[str] = ["Course", "Course Title", "SP20", "SU20", "FA20", "SP21", "SU21", "FA21"]
        cutoff: str = "FA20"
        expectation: dict[int, str] = {
            "FA20": 4,
            "SP21": 5,
            "SU21": 6,
            "FA21": 7
        }
        self.assertDictEqual(expectation, populate_column_semester_map(row, cutoff))
    
    def test_extract_data(self):
        # this uses a cut down version of the course schedule in the test resources directory
        output: dict[str, list[str]] = extract_sheet_data(load_workbook(self.resource_path / "schedule_input_test.xlsx").active)
        self.assertListEqual(output["CPSC 1105"], ["SP25", "SU25", "FA25", "SP26", "SU26", "FA26", "SP27", "SU27", "FA27", "SP28", "SU28", "FA28", "SP29"]) # all present
        self.assertListEqual(output["CPSC 1555"], []) # none present (blanks)
        self.assertListEqual(output["CPSC 2105"], ["SP25", "FA25", "SP26", "FA26", "SP27", "FA27", "SP28", "FA28", "SP29"]) #  no summers, also has blanks  indicated by space or period
        self.assertListEqual(output["CPSC 2108"], ["SP25", "SU25", "FA25", "SP26", "SU26", "FA26", "SP27", "SU27", "FA27", "SP28", "SU28", "FA28", "SP29"]) # all, some have only one letter
        self.assertListEqual(output["CYBR 3108"], ["SP25", "SP26", "SP27", "SP28", "SP29"]) # spring only, also has ?? in one cell with a valid O
        self.assertListEqual(output["CPSC 3137"], ["FA26", "FA28"]) # has extra text past last column, should be ignored
        self.assertListEqual(output["CPSC 3165"], ["SP25", "SU25", "FA25", "SP26", "SU26", "FA26", "SP27", "SU27", "FA27", "SP28", "SU28", "FA28", "SP29"]) # all present, some have (May)
        self.assertEqual(7, len(output.keys()))
    
    def test_extract_data_with_cutoff(self):
        # this uses a cut down version of the course schedule in the test resources directory
        # similar to test above but with a cutoff, so all results before SU27 (non inclusive) should be skipped
        output: dict[str, list[str]] = extract_sheet_data(load_workbook(self.resource_path / "schedule_input_test.xlsx").active, cutoff="SU27")
        self.assertListEqual(output["CPSC 1105"], ["SU27", "FA27", "SP28", "SU28", "FA28", "SP29"])
        self.assertListEqual(output["CPSC 1555"], [])
        self.assertListEqual(output["CPSC 2105"], ["FA27", "SP28", "FA28", "SP29"])
        self.assertListEqual(output["CPSC 2108"], ["SU27", "FA27", "SP28", "SU28", "FA28", "SP29"])
        self.assertListEqual(output["CYBR 3108"], ["SP28", "SP29"])
        self.assertListEqual(output["CPSC 3137"], ["FA28"])
        self.assertListEqual(output["CPSC 3165"], ["SU27", "FA27", "SP28", "SU28", "FA28", "SP29"])
        self.assertEqual(7, len(output.keys()))
