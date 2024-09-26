import unittest
from os import remove
from collections import OrderedDict

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from class_planning_tool.output_generation.class_plan_writer import write_plan_workbook

class TestClassScheduleParsing(unittest.TestCase):

    def setUp(self):
        self.sample_data: OrderedDict[str, list[dict[str, str]]] = OrderedDict()
        self.sample_data["Fall 2025"] = [
            {
                "code": "CPSC1111",
                "title": "Some Class 1",
            },
            {
                "code": "CPSC2222",
                "title": "Some Class 2",
            },
            {
                "code": "CPSC3333",
                "title": "Some Class 3",
            }
        ]
        self.sample_data["Spring 2026"] = []
        self.sample_data["Summer 2026"] = [
            {
                "code": "CPSC7777",
                "title": "Some Class 1",
            },
            {
                "code": "CPSC8888",
                "title": "Some Class 2",
            }
        ]
    

    def test_output_sheet_values(self):
        expected_values: dict[str, str] = {
            "B2": "Study Plan",

            "B3": "Fall 2025",
            "B4": "CPSC1111",
            "B5": "CPSC2222",
            "B6": "CPSC3333",
            "C4": "Some Class 1",
            "C5": "Some Class 2",
            "C6": "Some Class 3",

            "D3": "Spring 2026",
            "D4": None,
            "D5": None,
            "D6": None,
            "E4": None,
            "E5": None,
            "E6": None,

            "F3": "Summer 2026",
            "F4": "CPSC7777",
            "F5": "CPSC8888",
            "F6": None,
            "G4": "Some Class 1",
            "G5": "Some Class 2",
            "G6": None,

            "B8": "Courses: 3",
            "D8": "Courses: 0",
            "F8": "Courses: 2"
        }
        write_plan_workbook(self.sample_data, "Test_Plan.xlsx")
        ws: Worksheet = load_workbook("Test_Plan.xlsx").active
        for key, val in expected_values.items():
            self.assertEqual(val, ws[key].value)
        remove("Test_Plan.xlsx")